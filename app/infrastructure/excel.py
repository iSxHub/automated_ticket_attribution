from __future__ import annotations
import logging
from io import BytesIO
from typing import Iterable, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from app.domain.helpdesk import HelpdeskRequest
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)

class ExcelReportError(RuntimeError):
    """Raised when the Excel report cannot be generated."""

def build_excel(requests: Iterable[HelpdeskRequest]) -> bytes:

    # sorts by request_category, request_type, short_description (ascending).
    sorted_requests: List[HelpdeskRequest] = sorted(
        requests,
        key=lambda r: (
            (r.request_category or "").lower(),
            (r.request_type or "").lower(),
            (r.short_description or "").lower(),
        ),
    )

    # create workbook and sheet
    try:
        wb = Workbook()

        ws_raw = wb.active
        if ws_raw is None or not isinstance(ws_raw, Worksheet):
            logger.error("Active sheet is not a Worksheet or is None: %r", ws_raw)
            raise ExcelReportError("Failed to get active worksheet")
        ws: Worksheet = ws_raw

        ws.title = "Helpdesk Requests"

        # styles
        header_font = Font(bold=True, size=14)
        default_font = Font(size=14)
        header_fill = PatternFill(fill_type="solid", fgColor="FFC000")
        border_side = Side(border_style="medium", color="000000")
        default_border = Border(
            left=border_side,
            right=border_side,
            top=border_side,
            bottom=border_side,
        )

        # define and write header
        headers = [
            "raw_id",
            "request_category",
            "request_type",
            "short_description",
            "sla_value",
            "sla_unit",
        ]
        ws.append(headers)

        # write data rows
        for req in sorted_requests:
            ws.append(
                [
                    req.raw_id or "",
                    req.request_category or "",
                    req.request_type or "",
                    req.short_description or "",
                    req.sla_value if req.sla_value is not None else "",
                    req.sla_unit or "",
                ]
            )

        # apply styles
        for row_idx, row in enumerate(
                ws.iter_rows(
                    min_row=1,
                    max_row=ws.max_row,
                    max_col=ws.max_column,
                ),
                start=1,
        ):
            for cell in row:
                cell.border = default_border
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = default_font

        # auto-fit by setting column width from max content length
        for column_cells in ws.columns:
            max_length = 0

            first_cell = column_cells[0]
            col_index: Any = first_cell.column
            if not isinstance(col_index, int):
                # skip weird column indices instead of crashing
                logger.warning("Unexpected column index type: %r (%r)", col_index, type(col_index))
                continue

            column_letter = get_column_letter(col_index)
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                value_str = str(value)
                if len(value_str) > max_length:
                    max_length = len(value_str)

            # padding
            ws.column_dimensions[column_letter].width = max_length + 2

        # save workbook into an in-memory buffer and return bytes
        with BytesIO() as buffer:
            wb.save(buffer)
            return buffer.getvalue()

    except Exception as exc:
        logger.exception("Failed to build Excel report")
        raise ExcelReportError("Failed to build Excel report") from exc